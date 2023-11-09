import re
import openpyxl as wb
from models import db
from models import Signals
from datetime import datetime
from general_functions import General_functions as GF
from enum import Enum
import traceback

SHIFT = 1
LIST_MODULE = ['CPU', 'PSU', 'CN', 'MN', 'AI', 'AO', 'DI', 'RS', 'DO']

today = datetime.now()


class NameColumn(Enum):
    '''Перечисление статических столбцов таблицы.'''
    ID = 'id'
    TYPE_SIGNAL = 'type_signal'
    TAG = 'tag'
    NAME = 'description'
    SCHEMA = 'schema'
    KLK = 'klk'
    CONTACT = 'contact'
    BASKET = 'basket'
    MODULE = 'module'
    CHANNEl = 'channel'


class DataExel():
    '''Инициализация файла Exel и его таблиц.
    выдача массива для записи в базу SQL'''
    def __init__(self, exel: str, logtext):
        self.exel = exel
        self.logsTextEdit = logtext
        self.connect = wb.load_workbook(self.exel,
                                        read_only=True,
                                        data_only=True)

    def disconnect_exel(self):
        '''Разрыв связи с Exel.'''
        self.connect.close()

    def read_table(self) -> list:
        '''Список таблиц Exel.'''
        return [sheet.title for sheet in self.connect.worksheets]

    def max_column(self, uso: str):
        """Читаем выбранную таблицу и получаем макс-ное кол-во столбцов.

        Args:
            uso (str): название шкафа
        """
        self.sheet = self.connect[uso]
        return self.sheet.max_column

    def read_sheet(self, row: int, column: int) -> str:
        '''Чтение ячейки таблицы Exel.'''
        return self.sheet.cell(row=row, column=column).value

    def read_hat_table(self, uso: str, number_row: int,
                       is_tuple: bool,
                       select_column: tuple = None) -> dict | tuple:
        """Список с названиями столбцов.

        Args:
            uso (str): название шкафа
            number_row (int): номер строки с с названиями столбцов
            is_tuple (bool): флаг использования массива tuple
            select_column (tuple, optional): выбранные столбцы с
                                            номерами позиций.

        Returns:
            dict|tuple: либо то либо то
        """
        column = self.max_column(uso)

        if is_tuple:
            hat_tabl = {}
        else:
            hat_tabl = []

        for i in range(int(number_row), int(number_row) + SHIFT):
            for j in range(1, column + SHIFT):
                cell = self.read_sheet(i, j)
                if cell is None:
                    continue

                if is_tuple:
                    for key, value in select_column.items():
                        if value == cell:
                            hat_tabl[key] = j - SHIFT
                else:
                    hat_tabl.append(cell)

        return hat_tabl

    def database_count_row(self):
        '''Вычисляем кол-во строк в таблице Signals в базе SQL.
        Для дальнейшего добавления сигналов в конец таблицы'''
        cursor = db.cursor()
        try:
            cursor.execute('''SELECT COUNT(*) FROM signals''')
            count_row = cursor.fetchall()[0][0]
        except Exception:
            count_row = 0
        return count_row

    def search_type(self, scheme, type_signal):
        '''Дополнительная проверка на тип сигнала.'''
        for value in LIST_MODULE:
            if str(scheme).find(value) != -1:
                type_signal = value
                return type_signal
        return type_signal

    def sub_str(self, uso, basket, module, channel):
        '''Добавляем теги к резервам.'''
        dop_func = GF()
        tag = re.sub(r'(МНС)|(ПТ)|(САР)|(РП)|(БРУ)|\)|(c)', '', uso)
        tag = re.sub(r'(УСО)', 'USO', tag)
        tag = re.sub(r'\(|\.', '_', tag)
        tag = dop_func.translate(tag)
        tag = tag.replace(' ', '')
        tag = f'REZ{tag}_{basket}_{module}_{channel}'
        return tag

    def preparation_import(self, uso: str, number_row: int,
                           select_col: tuple) -> list:
        '''Подготовка таблицы к импорту.'''
        data = []

        count_row = self.database_count_row()

        tuple_name = self.read_hat_table(uso, number_row, True, select_col)

        for row in self.sheet.iter_rows(min_row=(int(number_row) + 1)):
            name = row[tuple_name[NameColumn.NAME.value]].value
            tag = row[tuple_name[NameColumn.TAG.value]].value
            type_s = row[tuple_name[NameColumn.TYPE_SIGNAL.value]].value
            schema = row[tuple_name[NameColumn.SCHEMA.value]].value
            klk = row[tuple_name[NameColumn.KLK.value]].value
            contact = row[tuple_name[NameColumn.CONTACT.value]].value
            basket = row[tuple_name[NameColumn.BASKET.value]].value
            module = row[tuple_name[NameColumn.MODULE.value]].value
            channel = row[tuple_name[NameColumn.CHANNEl.value]].value

            if (basket or module or channel) is None:
                continue
            count_row += 1

            if tag is None and 'RS' not in schema:
                tag = self.sub_str(uso, basket, module, channel)

            type_s = self.search_type(schema, type_s)

            data.append(dict(id=count_row,
                             type_signal=type_s,
                             uso=uso,
                             tag=tag,
                             description=name,
                             schema=schema,
                             klk=klk,
                             contact=contact,
                             basket=basket,
                             module=module,
                             channel=channel))
        return data


class Import_in_SQL(DataExel):
    '''Запись и обновление сигналов в базе SQL.'''
    def exists_signals(self, row: object, uso: str):
        '''Проверяем существование сигнала в базе
        по корзине, модулю, каналу.'''
        exist_row = Signals.select().where(Signals.uso == uso,
                                           Signals.basket == str(row[NameColumn.BASKET.value]),
                                           Signals.module == str(row[NameColumn.MODULE.value]),
                                           Signals.channel == str(row[NameColumn.CHANNEl.value]))
        return exist_row

    def compare_row(self, row_exel: dict, msg: str,
                    object_sql: object, object_exel: str) -> str:
        """Сравнение значений строки из базы SQL и таблицы Exel.

        Args:
            row_exel (dict): строка сигнала из Exel
            msg (str): сообщении о событии
            object_sql (object): запрос из базы SQL
            object_exel (str): имя столбца Exel

        Returns:
            str: сообщение
        """
        if str(object_sql) != str(row_exel[object_exel]):
            msg = f'{msg}{object_exel}: {row_exel[object_exel]}({object_sql}),'

        return msg

    def record_row(self, row_exel: dict, req_sql: object) -> str:
        """Обновление значения в строке сигнала.
        Формирование корректного сообщения об изменении.

        Args:
            row_exel (dict): строка сигнала из Exel
            req_sql (object): запрос из базы SQl

        Returns:
            str: cсообщение
        """
        dop_msg = ''
        for row in req_sql:
            dop_msg = self.compare_row(row_exel, dop_msg, row.tag,
                                       NameColumn.TAG.value)
            dop_msg = self.compare_row(row_exel, dop_msg, row.description,
                                       NameColumn.NAME.value)
            dop_msg = self.compare_row(row_exel, dop_msg, row.schema,
                                       NameColumn.SCHEMA.value)
            dop_msg = self.compare_row(row_exel, dop_msg, row.klk,
                                       NameColumn.KLK.value)
            dop_msg = self.compare_row(row_exel, dop_msg, row.contact,
                                       NameColumn.CONTACT.value)
            if dop_msg != '':
                Signals.update(**{NameColumn.TAG.value: row_exel[NameColumn.TAG.value],
                                  NameColumn.NAME.value: row_exel[NameColumn.NAME.value],
                                  NameColumn.SCHEMA.value: row_exel[NameColumn.SCHEMA.value],
                                  NameColumn.KLK.value: row_exel[NameColumn.KLK.value],
                                  NameColumn.CONTACT.value: row_exel[NameColumn.CONTACT.value]}).where(
                    Signals.id == row.id).execute()

                self.logsTextEdit.logs_msg(f'''Импорт КД:
                                           name = {row_exel[NameColumn.NAME.value]},
                                           id = {row.id}, {dop_msg} сигнал обновлен''', 0)

    def database_entry_SQL(self, data: dict, uso: str):
        '''По кнопке добавить новое УСО.
        Запись новых строк в базу SQL.'''
        with db.atomic():
            try:
                Signals.insert_many(data).execute()
                self.logsTextEdit.logs_msg(f'''Импорт КД: в таблицу signals
                                           добавлено новое УСО {uso}''', 1)
            except Exception:
                self.logsTextEdit.logs_msg(f'''Импорт КД: ошибка импорта
                                           {traceback.format_exc()}''', 2)

    def row_update_SQL(self, data: dict, uso: str):
        '''По кнопке обновить данные.
        Обновление старой записи, если имеется или запись новой строки.'''
        with db.atomic():
            try:
                for row_exel in data:

                    exists_s = self.exists_signals(row_exel, uso)

                    if not bool(exists_s):

                        Signals.create(**row_exel)
                        self.logsTextEdit.logs_msg(f'''Импорт КД:
                                          добавлен новый сигнал:
                                          id - {row_exel[NameColumn.ID.value]},
                                          description - {row_exel[NameColumn.NAME.value]},
                                          module - {row_exel[NameColumn.MODULE.value]},
                                          channel - {row_exel[NameColumn.CHANNEl.value]}''', 0)
                        continue
                    else:
                        messages = self.record_row(row_exel, exists_s)
                        if messages is not None:
                            self.logsTextEdit.logs_msg(messages, 3)
            except Exception:
                self.logsTextEdit.logs_msg(f'''Импорт КД:
                                           ошибка импорта
                                           {traceback.format_exc()}''', 2)
        self.logsTextEdit.logs_msg('Импорт КД: обновление завершен', 0)